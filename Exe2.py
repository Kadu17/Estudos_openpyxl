from openpyxl import load_workbook

wb = load_workbook('C:/Users/50947025804/PycharmProjects/Openpyxl/teste.xlsx')
p = wb['P1']
print(p.cell(row=1, column=2).value)